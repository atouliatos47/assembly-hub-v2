import re
import os
import json
import uuid
import socket
from datetime import datetime
from functools import wraps
from flask import Flask, request, jsonify, send_from_directory, session, redirect, url_for
from flask_sock import Sock
from openpyxl import load_workbook
from docx import Document

app = Flask(__name__, static_folder='public')
sock = Sock(app)

# ─── Auth config ───────────────────────────────────────────────────────────
# Set these as environment variables in Render dashboard for security
# Defaults are used for local development only
app.secret_key = os.environ.get('SECRET_KEY', 'assembly-hub-dev-secret-change-me')
DASHBOARD_PASSWORD = os.environ.get('DASHBOARD_PASSWORD', 'assembly2024')

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ─── In-memory state ───────────────────────────────────────────────────────
centers   = {}    # { id: { id, name, color, assignedDoc, currentPage, connected } }
documents = {}    # { id: { id, name, type, pages, uploadedAt, filePath } }
clients   = {}    # { id: { ws, role, centerId } }

# ─── Helpers ───────────────────────────────────────────────────────────────
def broadcast(msg):
    dead = []
    for cid, c in clients.items():
        try:
            c['ws'].send(json.dumps(msg))
        except Exception:
            dead.append(cid)
    for cid in dead:
        clients.pop(cid, None)

def notify_center(center_id, msg):
    dead = []
    for cid, c in clients.items():
        if c.get('centerId') == center_id:
            try:
                c['ws'].send(json.dumps(msg))
            except Exception:
                dead.append(cid)
    for cid in dead:
        clients.pop(cid, None)

def doc_summary(doc):
    return {
        'id': doc['id'],
        'name': doc['name'],
        'type': doc['type'],
        'pageCount': len(doc.get('pages', [])),
        'uploadedAt': doc['uploadedAt']
    }

# ─── File Parsing ──────────────────────────────────────────────────────────
def parse_excel(filepath):
    wb = load_workbook(filepath, data_only=True)
    pages = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_html = '<table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;width:100%;font-size:14px;">'
        for row in ws.iter_rows():
            rows_html += '<tr>'
            for cell in row:
                val = cell.value if cell.value is not None else ''
                bold = 'font-weight:bold;' if cell.font and cell.font.bold else ''
                bg = ''
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == 'rgb':
                    rgb = cell.fill.fgColor.rgb
                    if rgb and rgb != '00000000' and rgb != 'FFFFFFFF':
                        bg = f'background-color:#{rgb[2:]};'
                rows_html += f'<td style="{bold}{bg}padding:4px 8px;">{val}</td>'
            rows_html += '</tr>'
        rows_html += '</table>'
        pages.append({'title': sheet_name, 'html': rows_html})
    return pages

def parse_word(filepath):
    doc = Document(filepath)
    pages = []
    current_title = 'Page 1'
    current_html = ''
    page_num = 1

    for para in doc.paragraphs:
        style = para.style.name if para.style else ''
        text = para.text.strip()

        if style.startswith('Heading 1') or style.startswith('Heading 2'):
            # Save previous page if it has content
            if current_html.strip():
                pages.append({'title': current_title, 'html': current_html})
            current_title = text or f'Section {page_num}'
            tag = 'h1' if '1' in style else 'h2'
            current_html = f'<{tag}>{text}</{tag}>'
            page_num += 1
        else:
            if text:
                current_html += f'<p>{para.text}</p>'
            else:
                current_html += '<br>'

    if current_html.strip():
        pages.append({'title': current_title, 'html': current_html})

    if not pages:
        pages.append({'title': 'Document', 'html': '<p>Empty document</p>'})

    return pages

def make_slug(name):
    """Convert 'Assembly 1' → 'assembly-1', ensure uniqueness"""
    slug = name.lower().strip()
    slug = re.sub(r'[^a-z0-9]+', '-', slug).strip('-')
    base = slug
    counter = 2
    while any(c['slug'] == slug for c in centers.values()):
        slug = f"{base}-{counter}"
        counter += 1
    return slug


# ─── Auth helpers ──────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated

def api_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return jsonify({'error': 'Unauthorised'}), 401
        return f(*args, **kwargs)
    return decorated

# ─── Login routes ──────────────────────────────────────────────────────────
@app.route('/login', methods=['GET'])
def login_page():
    return send_from_directory('public', 'login.html')

@app.route('/login', methods=['POST'])
def do_login():
    data = request.json or {}
    if data.get('password') == DASHBOARD_PASSWORD:
        session['logged_in'] = True
        session.permanent = True
        return jsonify({'success': True})
    return jsonify({'error': 'Incorrect password'}), 401

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

@app.route('/')
@login_required
def index():
    return send_from_directory('public/dashboard', 'index.html')

@app.route('/dashboard')
@login_required
def dashboard():
    return send_from_directory('public/dashboard', 'index.html')

@app.route('/display/<center_ref>')
def display(center_ref):
    return send_from_directory('public/display', 'index.html')

@app.route('/public/<path:filename>')
def static_files(filename):
    return send_from_directory('public', filename)

@app.route('/dashboard/<path:filename>')
@login_required
def dashboard_static(filename):
    return send_from_directory('public/dashboard', filename)

@app.route('/sw.js')
def service_worker():
    return send_from_directory('public', 'sw.js', mimetype='application/javascript')

@app.route('/icons/<path:filename>')
def icons(filename):
    return send_from_directory('public/icons', filename)

@app.route('/display/manifest.json')
def display_manifest():
    return send_from_directory('public/display', 'manifest.json')


# ─── Documents API ─────────────────────────────────────────────────────────
@app.route('/api/documents', methods=['GET'])
@api_login_required
def get_documents():
    return jsonify([doc_summary(d) for d in documents.values()])

@app.route('/api/documents/upload', methods=['POST'])
@api_login_required
def upload_document():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    filename = file.filename
    ext = os.path.splitext(filename)[1].lower()

    if ext not in ['.xlsx', '.xls', '.docx', '.doc']:
        return jsonify({'error': 'Only Excel and Word files allowed'}), 400

    saved_name = f"{int(datetime.now().timestamp())}_{filename}"
    filepath = os.path.join(UPLOAD_DIR, saved_name)
    file.save(filepath)

    try:
        if ext in ['.xlsx', '.xls']:
            pages = parse_excel(filepath)
            doc_type = 'excel'
        else:
            pages = parse_word(filepath)
            doc_type = 'word'
    except Exception as e:
        os.remove(filepath)
        return jsonify({'error': f'Failed to parse file: {str(e)}'}), 500

    doc_name = request.form.get('name', filename)
    doc = {
        'id': str(uuid.uuid4()),
        'name': doc_name,
        'originalName': filename,
        'type': doc_type,
        'pages': pages,
        'uploadedAt': datetime.now().isoformat(),
        'filePath': filepath
    }
    documents[doc['id']] = doc
    broadcast({'type': 'DOCUMENT_ADDED', 'document': doc_summary(doc)})
    return jsonify({'success': True, 'document': doc_summary(doc)})

@app.route('/api/documents/<doc_id>/full', methods=['GET'])
@api_login_required
def get_document_full(doc_id):
    doc = documents.get(doc_id)
    if not doc:
        return jsonify({'error': 'Not found'}), 404
    return jsonify(doc)

@app.route('/api/documents/<doc_id>', methods=['DELETE'])
@api_login_required
def delete_document(doc_id):
    doc = documents.get(doc_id)
    if not doc:
        return jsonify({'error': 'Not found'}), 404

    for center in centers.values():
        if center.get('assignedDoc') and center['assignedDoc']['id'] == doc_id:
            center['assignedDoc'] = None
            center['currentPage'] = 0
            notify_center(center['id'], {'type': 'DOCUMENT_REMOVED'})
            broadcast({'type': 'CENTER_UPDATED', 'center': center})

    try:
        os.remove(doc['filePath'])
    except Exception:
        pass

    del documents[doc_id]
    broadcast({'type': 'DOCUMENT_DELETED', 'id': doc_id})
    return jsonify({'success': True})

# ─── Centers API ───────────────────────────────────────────────────────────
@app.route('/api/centers/resolve/<center_ref>', methods=['GET'])
def resolve_center(center_ref):
    # Try by slug first, then by ID
    center = next((c for c in centers.values() if c.get('slug') == center_ref), None)
    if not center:
        center = centers.get(center_ref)
    if not center:
        return jsonify({'error': 'Center not found'}), 404
    return jsonify({'id': center['id']})

@app.route('/api/centers', methods=['GET'])
@api_login_required
def get_centers():
    return jsonify(list(centers.values()))

@app.route('/api/centers', methods=['POST'])
@api_login_required
def create_center():
    data = request.json or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Name required'}), 400

    center = {
        'id': str(uuid.uuid4()),
        'name': name,
        'slug': make_slug(name),
        'color': data.get('color', '#2563eb'),
        'assignedDoc': None,
        'currentPage': 0,
        'connected': False,
        'createdAt': datetime.now().isoformat()
    }
    centers[center['id']] = center
    broadcast({'type': 'CENTER_ADDED', 'center': center})
    return jsonify({'success': True, 'center': center})

@app.route('/api/centers/<center_id>', methods=['PUT'])
@api_login_required
def update_center(center_id):
    center = centers.get(center_id)
    if not center:
        return jsonify({'error': 'Not found'}), 404
    data = request.json or {}
    if 'name' in data:
        center['name'] = data['name']
    if 'color' in data:
        center['color'] = data['color']
    broadcast({'type': 'CENTER_UPDATED', 'center': center})
    return jsonify({'success': True, 'center': center})

@app.route('/api/centers/<center_id>', methods=['DELETE'])
@api_login_required
def delete_center(center_id):
    if center_id not in centers:
        return jsonify({'error': 'Not found'}), 404
    del centers[center_id]
    broadcast({'type': 'CENTER_DELETED', 'id': center_id})
    return jsonify({'success': True})

@app.route('/api/centers/<center_id>/assign', methods=['POST'])
@api_login_required
def assign_document(center_id):
    center = centers.get(center_id)
    if not center:
        return jsonify({'error': 'Center not found'}), 404

    data = request.json or {}
    doc_id = data.get('documentId')

    if not doc_id:
        center['assignedDoc'] = None
        center['currentPage'] = 0
        notify_center(center_id, {'type': 'DOCUMENT_REMOVED'})
    else:
        doc = documents.get(doc_id)
        if not doc:
            return jsonify({'error': 'Document not found'}), 404
        center['assignedDoc'] = {'id': doc['id'], 'name': doc['name'], 'type': doc['type']}
        center['currentPage'] = 0
        notify_center(center_id, {'type': 'DOCUMENT_ASSIGNED', 'document': doc, 'currentPage': 0})

    broadcast({'type': 'CENTER_UPDATED', 'center': center})
    return jsonify({'success': True, 'center': center})

@app.route('/api/centers/<center_id>/page', methods=['POST'])
@api_login_required
def set_page(center_id):
    center = centers.get(center_id)
    if not center:
        return jsonify({'error': 'Center not found'}), 404
    if not center.get('assignedDoc'):
        return jsonify({'error': 'No document assigned'}), 400

    doc = documents.get(center['assignedDoc']['id'])
    if not doc:
        return jsonify({'error': 'Document not found'}), 404

    page = int((request.json or {}).get('page', 0))
    max_page = len(doc['pages']) - 1
    center['currentPage'] = max(0, min(page, max_page))

    notify_center(center_id, {
        'type': 'PAGE_CHANGE',
        'page': center['currentPage'],
        'totalPages': len(doc['pages'])
    })
    broadcast({'type': 'CENTER_UPDATED', 'center': center})
    return jsonify({'success': True, 'currentPage': center['currentPage']})

# ─── WebSocket ─────────────────────────────────────────────────────────────
@sock.route('/ws')
def websocket(ws):
    client_id = str(uuid.uuid4())
    clients[client_id] = {'ws': ws, 'role': None, 'centerId': None}

    # Send initial state
    try:
        ws.send(json.dumps({
            'type': 'INIT',
            'centers': list(centers.values()),
            'documents': [doc_summary(d) for d in documents.values()]
        }))
    except Exception:
        return

    try:
        while True:
            raw = ws.receive()
            if raw is None:
                break
            try:
                msg = json.loads(raw)
                handle_message(client_id, msg)
            except Exception as e:
                print(f'WS error: {e}')
    finally:
        client = clients.pop(client_id, None)
        if client and client.get('centerId'):
            cid = client['centerId']
            if cid in centers:
                centers[cid]['connected'] = False
                broadcast({'type': 'CENTER_UPDATED', 'center': centers[cid]})

def handle_message(client_id, msg):
    client = clients.get(client_id)
    if not client:
        return

    if msg.get('type') == 'REGISTER_DASHBOARD':
        client['role'] = 'dashboard'

    elif msg.get('type') == 'REGISTER_DISPLAY':
        client['role'] = 'display'
        center_id = msg.get('centerId')
        client['centerId'] = center_id

        center = centers.get(center_id)
        if center:
            center['connected'] = True
            broadcast({'type': 'CENTER_UPDATED', 'center': center})

            if center.get('assignedDoc'):
                doc = documents.get(center['assignedDoc']['id'])
                if doc:
                    try:
                        client['ws'].send(json.dumps({
                            'type': 'DOCUMENT_ASSIGNED',
                            'document': doc,
                            'currentPage': center['currentPage']
                        }))
                    except Exception:
                        pass

# ─── Start ─────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    PORT = int(os.environ.get('PORT', 8443))
    hostname = socket.gethostname()
    try:
        local_ip = socket.gethostbyname(hostname)
    except:
        local_ip = 'localhost'

    print('\n🏭 Assembly Hub Server Running')
    print(f'   Dashboard : http://{local_ip}:{PORT}/dashboard')
    print(f'   Local     : http://localhost:{PORT}/dashboard')
    print(f'   Network IP: {local_ip}\n')

    app.run(host='0.0.0.0', port=PORT, debug=False)
